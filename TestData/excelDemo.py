import openpyxl


book = openpyxl.load_workbook("C:\\Users\Karthic\\Desktop\\Chennai_RelocationEstimation.xlsx")
sheet = book.active
cell = sheet.cell(row=5, column=1)
print(cell.value)
sheet.cell(row=13, column=1).value = "Karthic"
print(sheet.cell(row=13, column=1).value)
print(sheet.max_row)
print(sheet.max_column)

